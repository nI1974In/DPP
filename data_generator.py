import random
from openpyxl import Workbook

military_branch = ['Сухопутные войска', 'Воздушно-космические войска', 'Военно-морской флот',
                   'Отдельные рода войск', 'Специальные войска']

categories = ['А', 'Б']


def generate_recruits(n: int = 100) -> None:
    surnames = ['Иванов', 'Петров', 'Сидоров', 'Козлов', 'Смирнов', 'Лопаткин', 'Монеткин', 'Лобов', 'Рылов',
                'Глинкин']

    names = ['Иван', 'Петр', 'Алексей', 'Дмитрий', 'Сергей', 'Арсений', 'Михаил', 'Денис', 'Егор', 'Владислав',
             'Максим']

    patronymics = ['Иванович', 'Петрович', 'Алексеевич', 'Дмитриевич', 'Сергеевич', 'Николаевич', 'Андреевич',
                   'Родионович']

    wb = Workbook()
    ws = wb.active

    ws['A1'] = 'ID'
    ws['B1'] = 'Фамилия'
    ws['C1'] = 'Имя'
    ws['D1'] = 'Отчество'
    ws['E1'] = 'Категория годности'
    ws['F1'] = 'Физическая оценка'
    ws['G1'] = 'Психологическая оценка'
    for i in range(len(military_branch)):
        ws[f"{chr(ord('H') + i)}1"] = f'Приоритет {i + 1}'

    def generate_conscript(recruit_id: int):
        surname = random.choice(surnames)
        firstname = random.choice(names)
        patronymic = random.choice(patronymics)
        category = random.choice(categories)
        physical = random.randint(10, 100)
        psychological = random.randint(10, 100)
        priority = random.sample(military_branch, len(military_branch))

        row = (recruit_id, surname, firstname, patronymic, category, physical, psychological, *priority)
        ws.append(row)

    for i in range(1, n + 1):
        generate_conscript(i)

    wb.save(r'res\conscript.xlsx')


def generate_military(recruiters_amount: int) -> None:
    wb = Workbook()
    ws = wb.active

    ws['A1'] = 'Род войск'
    ws['B1'] = 'Количество набираемых призывников'
    ws['C1'] = 'Минимальная категория годности'
    ws['D1'] = 'Минимальный балл физподготовки'
    ws['E1'] = 'Минимальный балл психологического отбора'

    counter = recruiters_amount
    for i in range(len(military_branch)):
        if i+1 < len(military_branch) and counter > 5:
            amount = 5 * random.randint(1, counter // 5)
            counter -= amount
        else:
            amount = counter
        category = random.choice(categories)
        physical = 5 * random.randint(2, 19)
        psychological = 5 * random.randint(2, 19)

        row = (military_branch[i], amount, category, physical, psychological)
        ws.append(row)

    wb.save(r'res\military.xlsx')


if __name__ == '__main__':
    generate_recruits(n=1000)
    generate_military(recruiters_amount=300)
